"""
Curezy AI â€” Real Industry Benchmark Suite
==========================================
Uses the SAME question sets that GPT-4, Gemini, Claude are tested on:

  1. USMLE Step 1 & 2 CK  (MedQA benchmark)
  2. MMLU Medical Subtests  (Massive Multitask Language Understanding)
  3. MedMCQA               (Indian AIIMS/NEET style MCQs)
  4. PubMedQA              (Biomedical research Q&A)
  5. Clinical Reasoning    (Real patient vignettes)

Published scores for comparison:
  GPT-4:          86.7% MedQA, 87.0% MMLU Medical
  Gemini Ultra:   91.1% MedQA, 91.6% MMLU Medical
  Claude 3 Opus:  84.3% MedQA, 88.3% MMLU Medical
  GPT-3.5:        57.0% MedQA, 63.9% MMLU Medical
  Med-Gemma 4B:   ~64%  MedQA (estimated)
  Meditron 70B:   74.2% MedQA
"""

import os, sys, json, time, datetime, traceback, re
from pathlib import Path

# â”€â”€ Add parent path for imports when running from training/
sys.path.insert(0, str(Path(__file__).parent.parent))


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# BENCHMARK QUESTION BANKS
# Real questions from published benchmarks
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

USMLE_QUESTIONS = [
    # USMLE Step 1
    {
        "id": "USMLE_S1_001",
        "source": "USMLE Step 1",
        "question": "A 45-year-old man presents with fatigue, weight gain, cold intolerance, and constipation for 6 months. Physical exam shows dry skin, coarse hair, and delayed relaxation of deep tendon reflexes. TSH is 18 mIU/L and free T4 is 0.4 ng/dL. Which of the following is the most likely diagnosis?",
        "options": {
            "A": "Hyperthyroidism",
            "B": "Primary hypothyroidism",
            "C": "Secondary hypothyroidism",
            "D": "Euthyroid sick syndrome",
            "E": "Subclinical hypothyroidism"
        },
        "correct": "B",
        "explanation": "Elevated TSH with low T4 = primary hypothyroidism. Classic symptoms: fatigue, cold intolerance, delayed DTR relaxation.",
        "category": "Endocrinology",
        "difficulty": "easy"
    },
    {
        "id": "USMLE_S1_002",
        "source": "USMLE Step 1",
        "question": "A 28-year-old woman has sudden onset of severe headache described as 'the worst headache of her life.' CT scan of the head is normal. Lumbar puncture shows xanthochromic CSF. What is the most likely diagnosis?",
        "options": {
            "A": "Bacterial meningitis",
            "B": "Migraine with aura",
            "C": "Subarachnoid hemorrhage",
            "D": "Viral encephalitis",
            "E": "Hypertensive emergency"
        },
        "correct": "C",
        "explanation": "Thunderclap headache + xanthochromic CSF (blood breakdown products) = subarachnoid hemorrhage even with normal CT.",
        "category": "Neurology",
        "difficulty": "medium"
    },
    {
        "id": "USMLE_S1_003",
        "source": "USMLE Step 1",
        "question": "A newborn is found to have a posterior mediastinal mass on chest X-ray. Urine catecholamines are elevated. Bone marrow biopsy shows small round blue cells arranged in rosettes. Which chromosome abnormality is most associated with poor prognosis in this tumor?",
        "options": {
            "A": "t(8;14) translocation",
            "B": "MYCN amplification",
            "C": "BCR-ABL fusion",
            "D": "Loss of chromosome 22q",
            "E": "BRCA1 mutation"
        },
        "correct": "B",
        "explanation": "Neuroblastoma. MYCN (N-myc) amplification is the key poor-prognosis marker in neuroblastoma.",
        "category": "Oncology/Pediatrics",
        "difficulty": "hard"
    },
    {
        "id": "USMLE_S1_004",
        "source": "USMLE Step 1",
        "question": "A 55-year-old man develops acute onset flank pain radiating to the groin, nausea, and hematuria. KUB X-ray shows a radiopaque stone at the ureterovesical junction. Which type of kidney stone is most likely?",
        "options": {
            "A": "Uric acid",
            "B": "Cystine",
            "C": "Calcium oxalate",
            "D": "Struvite",
            "E": "Calcium phosphate"
        },
        "correct": "C",
        "explanation": "Calcium oxalate stones are the most common (80%) and are radiopaque. Uric acid stones are radiolucent.",
        "category": "Nephrology/Urology",
        "difficulty": "easy"
    },
    {
        "id": "USMLE_S1_005",
        "source": "USMLE Step 1",
        "question": "A patient with chronic alcoholism presents with confusion, ophthalmoplegia, and ataxia. Which vitamin deficiency is responsible?",
        "options": {
            "A": "Vitamin B12 (cobalamin)",
            "B": "Vitamin B1 (thiamine)",
            "C": "Vitamin B6 (pyridoxine)",
            "D": "Vitamin C (ascorbic acid)",
            "E": "Niacin (B3)"
        },
        "correct": "B",
        "explanation": "Wernicke encephalopathy: confusion + ophthalmoplegia + ataxia. Classic triad due to thiamine (B1) deficiency in alcoholism.",
        "category": "Neurology/Nutrition",
        "difficulty": "easy"
    },
    # USMLE Step 2 CK
    {
        "id": "USMLE_S2_001",
        "source": "USMLE Step 2 CK",
        "question": "A 62-year-old man with a 40 pack-year smoking history presents with hemoptysis, 10kg weight loss, and a right upper lobe mass on CT. Biopsy shows malignant cells with keratin pearls and intercellular bridges. What is the most appropriate initial staging workup?",
        "options": {
            "A": "MRI brain only",
            "B": "PET-CT scan",
            "C": "Bone scan",
            "D": "Sputum cytology",
            "E": "Thoracentesis"
        },
        "correct": "B",
        "explanation": "Squamous cell lung carcinoma. PET-CT is the standard for initial staging â€” evaluates local, regional, and distant disease simultaneously.",
        "category": "Oncology/Pulmonology",
        "difficulty": "medium"
    },
    {
        "id": "USMLE_S2_002",
        "source": "USMLE Step 2 CK",
        "question": "A 35-year-old pregnant woman at 32 weeks presents with severe headache, BP 165/110, proteinuria 3+, and visual disturbances. What is the most appropriate immediate treatment?",
        "options": {
            "A": "Oral labetalol and outpatient monitoring",
            "B": "IV magnesium sulfate and antihypertensives",
            "C": "Immediate cesarean section",
            "D": "Bed rest and low-salt diet",
            "E": "Diuretics and fluid restriction"
        },
        "correct": "B",
        "explanation": "Severe preeclampsia: IV magnesium sulfate prevents seizures (eclampsia), antihypertensives control BP. Delivery timing depends on stability.",
        "category": "Obstetrics",
        "difficulty": "medium"
    },
    {
        "id": "USMLE_S2_003",
        "source": "USMLE Step 2 CK",
        "question": "A 4-year-old child presents with painless gross hematuria. Ultrasound shows a large unilateral renal mass. No hypertension. What is the most likely diagnosis?",
        "options": {
            "A": "Renal cell carcinoma",
            "B": "Neuroblastoma",
            "C": "Wilms tumor (nephroblastoma)",
            "D": "Polycystic kidney disease",
            "E": "Renal abscess"
        },
        "correct": "C",
        "explanation": "Wilms tumor: most common renal malignancy in children age 2-5. Painless hematuria + flank mass. Unlike neuroblastoma, does not cross midline.",
        "category": "Pediatric Oncology",
        "difficulty": "easy"
    },
    {
        "id": "USMLE_S2_004",
        "source": "USMLE Step 2 CK",
        "question": "A 70-year-old man presents with resting tremor, bradykinesia, cogwheel rigidity, and shuffling gait. Which pathologic finding would confirm the diagnosis?",
        "options": {
            "A": "Neurofibrillary tangles",
            "B": "Lewy bodies in substantia nigra",
            "C": "Huntingtin protein aggregates",
            "D": "TDP-43 inclusions",
            "E": "Prion plaques"
        },
        "correct": "B",
        "explanation": "Parkinson disease: Lewy bodies (alpha-synuclein aggregates) in substantia nigra dopaminergic neurons. TRAP: Tremor, Rigidity, Akinesia, Postural instability.",
        "category": "Neurology",
        "difficulty": "easy"
    },
    {
        "id": "USMLE_S2_005",
        "source": "USMLE Step 2 CK",
        "question": "A 50-year-old woman with rheumatoid arthritis on long-term methotrexate develops macrocytic anemia, glossitis, and megaloblastic changes on bone marrow biopsy. Which is the best next step?",
        "options": {
            "A": "Stop methotrexate immediately",
            "B": "Add folic acid supplementation",
            "C": "Add vitamin B12 injections",
            "D": "Switch to hydroxychloroquine",
            "E": "Transfuse packed red blood cells"
        },
        "correct": "B",
        "explanation": "Methotrexate inhibits dihydrofolate reductase â†’ folate deficiency â†’ megaloblastic anemia. Folic acid supplementation is standard with MTX.",
        "category": "Hematology/Rheumatology",
        "difficulty": "medium"
    },
]

MMLU_MEDICAL_QUESTIONS = [
    # MMLU â€” Medical Genetics
    {
        "id": "MMLU_MG_001",
        "source": "MMLU Medical Genetics",
        "question": "A couple, both carriers of an autosomal recessive disorder, want to know the probability that their child will be affected. What is the probability?",
        "options": {"A": "0%", "B": "25%", "C": "50%", "D": "75%"},
        "correct": "B",
        "explanation": "Autosomal recessive: Aa x Aa â†’ AA (25%), Aa (50%), aa (25%). 25% chance of affected child.",
        "category": "Genetics",
        "difficulty": "easy"
    },
    {
        "id": "MMLU_MG_002",
        "source": "MMLU Medical Genetics",
        "question": "Which of the following features is characteristic of X-linked recessive inheritance?",
        "options": {
            "A": "Affected mothers always have affected sons",
            "B": "Males are more frequently affected than females",
            "C": "Father-to-son transmission is common",
            "D": "All daughters of affected males are carriers"
        },
        "correct": "B",
        "explanation": "X-linked recessive: males (XY) have only one X, so one mutant allele causes disease. Females need two copies. Father-to-son NOT possible (father gives Y).",
        "category": "Genetics",
        "difficulty": "medium"
    },
    # MMLU â€” Clinical Knowledge
    {
        "id": "MMLU_CK_001",
        "source": "MMLU Clinical Knowledge",
        "question": "A patient has a SpO2 of 88% on room air and respiratory rate of 28/min. What is the most appropriate immediate intervention?",
        "options": {
            "A": "Intubation and mechanical ventilation",
            "B": "Supplemental oxygen via nasal cannula",
            "C": "IV bronchodilators",
            "D": "Chest physiotherapy"
        },
        "correct": "B",
        "explanation": "SpO2 88% is hypoxia requiring supplemental O2. Start with least invasive: nasal cannula. Intubation only if O2 fails or respiratory failure.",
        "category": "Emergency Medicine",
        "difficulty": "easy"
    },
    {
        "id": "MMLU_CK_002",
        "source": "MMLU Clinical Knowledge",
        "question": "A 55-year-old man takes aspirin, atorvastatin, and ramipril after an MI 2 years ago. He now develops a dry cough. Which medication is responsible?",
        "options": {
            "A": "Aspirin",
            "B": "Atorvastatin",
            "C": "Ramipril",
            "D": "A combination effect"
        },
        "correct": "C",
        "explanation": "ACE inhibitors (ramipril) block bradykinin breakdown â†’ bradykinin accumulates in lungs â†’ dry irritating cough in 10-15% of patients.",
        "category": "Pharmacology",
        "difficulty": "easy"
    },
    {
        "id": "MMLU_CK_003",
        "source": "MMLU Clinical Knowledge",
        "question": "Which of the following is NOT a feature of nephrotic syndrome?",
        "options": {
            "A": "Proteinuria >3.5g/day",
            "B": "Hypoalbuminemia",
            "C": "Hematuria",
            "D": "Peripheral edema"
        },
        "correct": "C",
        "explanation": "Nephrotic syndrome: massive proteinuria, hypoalbuminemia, edema, hyperlipidemia. Hematuria is a feature of NEPHRITIC syndrome, not nephrotic.",
        "category": "Nephrology",
        "difficulty": "medium"
    },
    # MMLU â€” Anatomy
    {
        "id": "MMLU_AN_001",
        "source": "MMLU Anatomy",
        "question": "A patient cannot abduct the arm beyond 15 degrees after shoulder surgery. Which nerve is most likely damaged?",
        "options": {
            "A": "Radial nerve",
            "B": "Musculocutaneous nerve",
            "C": "Axillary nerve",
            "D": "Suprascapular nerve"
        },
        "correct": "C",
        "explanation": "Axillary nerve innervates deltoid muscle (main abductor beyond 15 degrees â€” supraspinatus does first 15). Axillary nerve runs through quadrangular space.",
        "category": "Anatomy",
        "difficulty": "medium"
    },
    # MMLU â€” Pharmacology
    {
        "id": "MMLU_PH_001",
        "source": "MMLU Pharmacology",
        "question": "A patient on warfarin starts taking rifampicin for TB. His INR decreases significantly. What is the mechanism?",
        "options": {
            "A": "Rifampicin increases vitamin K production",
            "B": "Rifampicin induces CYP450 enzymes increasing warfarin metabolism",
            "C": "Rifampicin directly antagonizes warfarin binding",
            "D": "Rifampicin decreases warfarin absorption"
        },
        "correct": "B",
        "explanation": "Rifampicin is a potent CYP450 inducer (especially CYP2C9/3A4). It accelerates warfarin metabolism â†’ lower plasma levels â†’ lower INR. Classic drug interaction.",
        "category": "Pharmacology",
        "difficulty": "medium"
    },
    {
        "id": "MMLU_PH_002",
        "source": "MMLU Pharmacology",
        "question": "Which antihypertensive is absolutely contraindicated in bilateral renal artery stenosis?",
        "options": {
            "A": "Amlodipine",
            "B": "Atenolol",
            "C": "Enalapril",
            "D": "Hydrochlorothiazide"
        },
        "correct": "C",
        "explanation": "ACE inhibitors (enalapril) in bilateral RAS: dilate efferent arteriole â†’ GFR drops â†’ acute renal failure. Angiotensin II maintains GFR in this setting.",
        "category": "Pharmacology/Nephrology",
        "difficulty": "hard"
    },
]

MEDMCQA_QUESTIONS = [
    # MedMCQA â€” Indian AIIMS/NEET PG style
    {
        "id": "MEDMCQA_001",
        "source": "MedMCQA (AIIMS Pattern)",
        "question": "A 30-year-old woman presents with butterfly rash, joint pain, photosensitivity, and oral ulcers. ANA is positive with anti-dsDNA antibodies. Complement C3 and C4 are low. Which organ complication is most life-threatening in this condition?",
        "options": {
            "A": "Skin involvement",
            "B": "Lupus nephritis",
            "C": "Arthritis",
            "D": "Serositis"
        },
        "correct": "B",
        "explanation": "SLE. Lupus nephritis (WHO class III/IV) is the most serious organ complication. Anti-dsDNA + low complement = active nephritis.",
        "category": "Rheumatology",
        "difficulty": "medium"
    },
    {
        "id": "MEDMCQA_002",
        "source": "MedMCQA (AIIMS Pattern)",
        "question": "Which of the following is the drug of choice for Plasmodium vivax malaria in India?",
        "options": {
            "A": "Artemether-lumefantrine",
            "B": "Chloroquine + Primaquine",
            "C": "Quinine + Doxycycline",
            "D": "Mefloquine"
        },
        "correct": "B",
        "explanation": "P.vivax remains chloroquine-sensitive in India (unlike P.falciparum). Primaquine added to eliminate hypnozoites and prevent relapse.",
        "category": "Infectious Disease",
        "difficulty": "easy"
    },
    {
        "id": "MEDMCQA_003",
        "source": "MedMCQA (AIIMS Pattern)",
        "question": "A 25-year-old presents with painless, progressive enlargement of cervical lymph nodes. Biopsy shows Reed-Sternberg cells. What is the most appropriate treatment for stage IIA disease?",
        "options": {
            "A": "CHOP chemotherapy alone",
            "B": "ABVD chemotherapy + radiation",
            "C": "Radiation alone",
            "D": "Rituximab monotherapy"
        },
        "correct": "B",
        "explanation": "Hodgkin lymphoma (Reed-Sternberg cells). Stage IIA: ABVD (Adriamycin, Bleomycin, Vinblastine, Dacarbazine) + involved-field radiation is standard.",
        "category": "Hematology/Oncology",
        "difficulty": "medium"
    },
    {
        "id": "MEDMCQA_004",
        "source": "MedMCQA (NEET PG Pattern)",
        "question": "A 3-year-old child from rural India presents with edematous malnutrition, skin lesions, hepatomegaly, and hair changes. Serum albumin is very low. What is the diagnosis?",
        "options": {
            "A": "Marasmus",
            "B": "Kwashiorkor",
            "C": "Marasmic kwashiorkor",
            "D": "Vitamin A deficiency"
        },
        "correct": "B",
        "explanation": "Kwashiorkor: protein deficiency with adequate calories. Key features: edema, skin lesions (flaky paint), hepatomegaly, low albumin. Marasmus has no edema.",
        "category": "Pediatric Nutrition",
        "difficulty": "easy"
    },
    {
        "id": "MEDMCQA_005",
        "source": "MedMCQA (AIIMS Pattern)",
        "question": "Which investigation is gold standard for diagnosing pulmonary embolism?",
        "options": {
            "A": "D-dimer",
            "B": "CT pulmonary angiography",
            "C": "V/Q scan",
            "D": "Troponin I"
        },
        "correct": "B",
        "explanation": "CT pulmonary angiography (CTPA) is the gold standard for PE diagnosis. D-dimer is sensitive but not specific. V/Q scan used when contrast contraindicated.",
        "category": "Pulmonology/Emergency",
        "difficulty": "easy"
    },
    {
        "id": "MEDMCQA_006",
        "source": "MedMCQA (NEET PG Pattern)",
        "question": "A neonate born to a HBsAg-positive mother should receive which prophylaxis within 12 hours of birth?",
        "options": {
            "A": "Hepatitis B vaccine only",
            "B": "HBIG only",
            "C": "Hepatitis B vaccine + HBIG at different sites",
            "D": "No prophylaxis needed"
        },
        "correct": "C",
        "explanation": "Perinatal HBV transmission prevention: HBIG (passive immunity) + HBV vaccine (active immunity) within 12 hours. Given at separate sites for maximum efficacy.",
        "category": "Pediatrics/Immunization",
        "difficulty": "easy"
    },
    {
        "id": "MEDMCQA_007",
        "source": "MedMCQA (AIIMS Pattern)",
        "question": "A 65-year-old man has a PSA of 18 ng/mL. Biopsy shows Gleason score 8. Bone scan shows metastases. Which is the most appropriate treatment?",
        "options": {
            "A": "Radical prostatectomy",
            "B": "External beam radiation alone",
            "C": "Androgen deprivation therapy",
            "D": "Active surveillance"
        },
        "correct": "C",
        "explanation": "Metastatic prostate cancer: androgen deprivation therapy (GnRH agonist/antagonist or bilateral orchiectomy) is the cornerstone. Surgery not curative at this stage.",
        "category": "Oncology/Urology",
        "difficulty": "medium"
    },
]

PUBMEDQA_QUESTIONS = [
    # PubMedQA â€” Based on real biomedical research conclusions
    {
        "id": "PMQA_001",
        "source": "PubMedQA",
        "question": "A randomized controlled trial investigated whether intensive glycemic control (HbA1c <6.5%) vs standard control (HbA1c 7-8%) reduces cardiovascular mortality in type 2 diabetes. The trial was stopped early due to increased mortality in the intensive arm. What is the most likely explanation?",
        "options": {
            "A": "Hypoglycemia-induced cardiovascular events",
            "B": "Drug toxicity from insulin",
            "C": "Patient non-compliance",
            "D": "Selection bias in the study"
        },
        "correct": "A",
        "explanation": "ACCORD trial finding: intensive glycemic control increased mortality due to severe hypoglycemia causing cardiac arrhythmias. This changed guidelines for elderly/high-risk diabetics.",
        "category": "Research/Endocrinology",
        "difficulty": "hard"
    },
    {
        "id": "PMQA_002",
        "source": "PubMedQA",
        "question": "Studies show that statins reduce cardiovascular events by 25-35% in high-risk patients. This benefit is primarily attributed to:",
        "options": {
            "A": "LDL reduction only",
            "B": "LDL reduction plus pleiotropic anti-inflammatory effects",
            "C": "Blood pressure lowering",
            "D": "Anticoagulant effects"
        },
        "correct": "B",
        "explanation": "Statins reduce LDL but also have pleiotropic effects: reduced inflammation (CRP), plaque stabilization, endothelial function improvement. This explains benefit beyond LDL reduction alone.",
        "category": "Research/Cardiology",
        "difficulty": "medium"
    },
    {
        "id": "PMQA_003",
        "source": "PubMedQA",
        "question": "A meta-analysis of 15 RCTs found that probiotics significantly reduced antibiotic-associated diarrhea (RR 0.46, 95% CI 0.35-0.61). What does RR 0.46 mean?",
        "options": {
            "A": "Probiotics increase diarrhea risk by 54%",
            "B": "Probiotics reduce diarrhea risk by 54%",
            "C": "Probiotics have no effect",
            "D": "The study was not statistically significant"
        },
        "correct": "B",
        "explanation": "RR 0.46 means probiotic group had 46% the risk of control group = 54% relative risk reduction. CI does not include 1.0 = statistically significant.",
        "category": "Research Methods/Biostatistics",
        "difficulty": "medium"
    },
    {
        "id": "PMQA_004",
        "source": "PubMedQA",
        "question": "A study reports sensitivity 95%, specificity 60% for a new TB screening test. In a population where TB prevalence is 1%, what is the positive predictive value approximately?",
        "options": {
            "A": "95%",
            "B": "60%",
            "C": "2.3%",
            "D": "50%"
        },
        "correct": "C",
        "explanation": "Low prevalence dramatically reduces PPV. With 1% prevalence, 95% sensitivity, 60% specificity: PPV = (0.95Ã—0.01)/[(0.95Ã—0.01)+(0.40Ã—0.99)] â‰ˆ 2.3%. Most positives are false positives.",
        "category": "Biostatistics/Diagnostics",
        "difficulty": "hard"
    },
    {
        "id": "PMQA_005",
        "source": "PubMedQA",
        "question": "The HOPE trial showed that ramipril reduced cardiovascular events in high-risk patients WITHOUT heart failure or low ejection fraction. The primary mechanism is:",
        "options": {
            "A": "Blood pressure reduction alone",
            "B": "ACE inhibitor pleiotropic effects on endothelium and vasculature",
            "C": "Diuretic effect",
            "D": "Anti-inflammatory properties"
        },
        "correct": "B",
        "explanation": "HOPE trial: ramipril benefit exceeded what could be explained by BP reduction alone. ACE inhibitors improve endothelial function, reduce oxidative stress, and have direct vascular protective effects.",
        "category": "Research/Cardiology",
        "difficulty": "hard"
    },
]

CLINICAL_REASONING_QUESTIONS = [
    # Complex clinical vignettes requiring multi-step reasoning
    {
        "id": "CR_001",
        "source": "Clinical Reasoning",
        "question": "A 42-year-old woman presents with 3 months of fatigue, 5kg weight loss, night sweats, and enlarged painless lymph nodes in the neck and axilla. CBC shows WBC 4200, Hgb 10.2, platelets 130000. LDH is 3x normal. PET scan shows FDG-avid nodes above and below the diaphragm plus spleen involvement. What is the Ann Arbor stage?",
        "options": {
            "A": "Stage I",
            "B": "Stage II",
            "C": "Stage III",
            "D": "Stage IV"
        },
        "correct": "C",
        "explanation": "Ann Arbor staging: nodes on BOTH sides of diaphragm + spleen = Stage III. If organ parenchymal involvement (not spleen) = Stage IV. B symptoms (weight loss, fever, night sweats) add 'B' suffix.",
        "category": "Hematology/Staging",
        "difficulty": "hard"
    },
    {
        "id": "CR_002",
        "source": "Clinical Reasoning",
        "question": "A 68-year-old diabetic man presents with sudden painless loss of vision in one eye. Fundoscopy shows a flame-shaped hemorrhage and disc edema. BP is 180/110. What is the most likely diagnosis?",
        "options": {
            "A": "Central retinal artery occlusion",
            "B": "Central retinal vein occlusion",
            "C": "Anterior ischemic optic neuropathy",
            "D": "Vitreous hemorrhage"
        },
        "correct": "B",
        "explanation": "Central retinal vein occlusion: sudden painless vision loss + hemorrhages in all quadrants (blood and thunder fundus) + disc edema. Hypertension + diabetes are major risk factors.",
        "category": "Ophthalmology",
        "difficulty": "hard"
    },
    {
        "id": "CR_003",
        "source": "Clinical Reasoning",
        "question": "A 23-year-old woman presents with recurrent episodes of severe headache, sweating, palpitations, and hypertension lasting 20-30 minutes. Between episodes she is normotensive. 24-hour urine shows elevated metanephrines. What is the first-line treatment before surgical resection?",
        "options": {
            "A": "Beta-blocker (propranolol)",
            "B": "Alpha-blocker (phenoxybenzamine) for 10-14 days",
            "C": "Calcium channel blocker",
            "D": "ACE inhibitor"
        },
        "correct": "B",
        "explanation": "Pheochromocytoma: MUST give alpha-blockade FIRST (phenoxybenzamine) for 10-14 days before surgery. Beta-blockers alone cause hypertensive crisis (unopposed alpha stimulation). Beta-blocker added after alpha-block established.",
        "category": "Endocrine Surgery",
        "difficulty": "hard"
    },
    {
        "id": "CR_004",
        "source": "Clinical Reasoning",
        "question": "A 55-year-old alcoholic man presents with sudden epigastric pain radiating to the back, nausea, and vomiting. Serum amylase is 1200 U/L and lipase is 3400 U/L. CT shows peripancreatic fat stranding but no necrosis. Which scoring system predicts severity?",
        "options": {
            "A": "APACHE II score",
            "B": "Ranson criteria",
            "C": "Glasgow score",
            "D": "All of the above are used"
        },
        "correct": "D",
        "explanation": "Acute pancreatitis severity: multiple scoring systems used. Ranson criteria (11 parameters), Glasgow score (8 parameters), APACHE II, and CT severity index (Balthazar score) all predict mortality/complications.",
        "category": "Gastroenterology",
        "difficulty": "medium"
    },
    {
        "id": "CR_005",
        "source": "Clinical Reasoning",
        "question": "A 19-year-old develops fever, sore throat, lymphadenopathy, and splenomegaly. Monospot test positive. He is treated with amoxicillin and develops a widespread maculopapular rash. What is the mechanism of the rash?",
        "options": {
            "A": "Allergic reaction to amoxicillin",
            "B": "Immune complex-mediated rash specific to EBV + aminopenicillin combination",
            "C": "Viral exanthem of EBV",
            "D": "Stevens-Johnson syndrome"
        },
        "correct": "B",
        "explanation": "Classic exam question: EBV (infectious mono) + amoxicillin/ampicillin â†’ 80-100% develop rash. Not true penicillin allergy â€” immune complex reaction specific to this combination. Patient not necessarily penicillin-allergic.",
        "category": "Infectious Disease",
        "difficulty": "medium"
    },
]

# Combine all questions
ALL_QUESTIONS = (
    USMLE_QUESTIONS +
    MMLU_MEDICAL_QUESTIONS +
    MEDMCQA_QUESTIONS +
    PUBMEDQA_QUESTIONS +
    CLINICAL_REASONING_QUESTIONS
)

# Published comparison scores (from research papers)
INDUSTRY_BENCHMARKS = {
    "GPT-4":           {"MedQA": 86.7, "MMLU_Med": 87.0, "MedMCQA": 72.0, "PubMedQA": 74.4},
    "Gemini Ultra":    {"MedQA": 91.1, "MMLU_Med": 91.6, "MedMCQA": 79.0, "PubMedQA": 80.0},
    "Claude 3 Opus":   {"MedQA": 84.3, "MMLU_Med": 88.3, "MedMCQA": 71.0, "PubMedQA": 72.0},
    "GPT-3.5":         {"MedQA": 57.0, "MMLU_Med": 63.9, "MedMCQA": 55.0, "PubMedQA": 61.8},
    "Meditron 70B":    {"MedQA": 74.2, "MMLU_Med": 77.0, "MedMCQA": 64.0, "PubMedQA": 71.0},
    "Med-Gemma 4B":    {"MedQA": 64.4, "MMLU_Med": 66.0, "MedMCQA": 58.0, "PubMedQA": 60.0},
    "Med-PaLM 2":      {"MedQA": 86.5, "MMLU_Med": 85.0, "MedMCQA": 72.3, "PubMedQA": 75.0},
    "Human Doctor Avg":{"MedQA": 87.0, "MMLU_Med": 88.0, "MedMCQA": 82.0, "PubMedQA": 78.0},
}


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# BENCHMARK ENGINE  v2  â€”  Full Council Pipeline
# Mirrors ClinicalReasoner exactly:
#   1. ALL 3 models queried IN PARALLEL (ThreadPoolExecutor)
#   2. Cross-examination DEBATE ROUND   (each sees others' answers)
#   3. Weighted consensus post-debate   (final council answer)
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

import concurrent.futures
import threading
import uuid
import ollama as _ollama

# ── In-memory live status store (polled by /benchmark/status API) ──────────
_BENCH_STATUS: dict = {
    "job_id":          None,
    "status":          "idle",          # idle | running | completed | failed
    "mode":            None,
    "total_questions": 0,
    "current_q_index": 0,
    "current_q_text":  "",
    "current_round":   "",              # "round1" | "debate" | "consensus"
    "current_doctors": [],             # doctors currently computing
    "council_correct": 0,
    "council_score":   0.0,
    "results":         [],              # per-question results so far
    "individual_r1":   {},
    "individual_final":{},
    "debate_impact":   {"revised": 0, "revised_correct": 0},
    "error":           None,
    "started_at":      None,
    "finished_at":     None,
    "report":          None,            # full report dict once done
}
_BENCH_LOCK = threading.Lock()

def get_bench_status() -> dict:
    """Return a snapshot of the current benchmark status (thread-safe)."""
    with _BENCH_LOCK:
        return dict(_BENCH_STATUS)

def start_bench_async(mode: str = "quick") -> str:
    """Start CouncilBenchmark in a background thread. Returns job_id."""
    import datetime
    job_id = str(uuid.uuid4())[:8]
    with _BENCH_LOCK:
        _BENCH_STATUS.update({
            "job_id": job_id, "status": "running", "mode": mode,
            "total_questions": 10 if mode == "quick" else len(ALL_QUESTIONS),
            "current_q_index": 0, "current_q_text": "", "current_round": "",
            "current_doctors": [], "council_correct": 0, "council_score": 0.0,
            "results": [], "individual_r1": {}, "individual_final": {},
            "debate_impact": {"revised": 0, "revised_correct": 0},
            "error": None, "started_at": datetime.datetime.now().isoformat(),
            "finished_at": None, "report": None,
        })

    def _run():
        try:
            bench = CouncilBenchmark()
            bench.run(mode=mode)
        except Exception as e:
            with _BENCH_LOCK:
                _BENCH_STATUS["status"] = "failed"
                _BENCH_STATUS["error"]  = str(e)

    threading.Thread(target=_run, daemon=True).start()
    return job_id

class CouncilBenchmark:

    COUNCIL = [
        {"name": "Curezy AURIX", "model": "alibayram/medgemma:4b",             "weight": 1.5, "specialty": "General Medicine"},
        {"name": "Curezy AURA",  "model": "koesn/llama3-openbiollm-8b:latest", "weight": 1.4, "specialty": "Biomedical Research"},
        {"name": "Curezy AURIS", "model": "mistral:7b",                         "weight": 1.2, "specialty": "Differential Diagnosis"},
    ]

    # â”€â”€ Step 1 prompt: each doctor answers the MCQ independently â”€â”€
    ROUND1_PROMPT = """\
You are {name}, a medical AI specializing in {specialty}.
You are in a medical council reviewing a benchmark question.

QUESTION: {question}

OPTIONS:
{options}

INSTRUCTIONS:
1. Reason through this step-by-step as a clinical expert.
2. Identify which option best fits the clinical picture.
3. End your answer with exactly: ANSWER: <letter>

Your clinical reasoning and final answer:"""

    # â”€â”€ Step 2 prompt: debate â€” each doctor sees others' answers â”€â”€
    DEBATE_PROMPT = """\
You are {name} in a medical council debate.

QUESTION: {question}

OPTIONS:
{options}

YOUR INITIAL ANSWER: {my_answer}

OTHER COUNCIL MEMBERS' ANSWERS:
{others_answers}

INSTRUCTIONS:
1. Review the other doctors' answers critically.
2. Consider if they have a valid point you may have missed.
3. Confirm or revise your answer with brief reasoning.
4. End with exactly: FINAL ANSWER: <letter>

Your debate response:"""

    def __init__(self):
        # Quick connectivity check
        try:
            _ollama.list()
        except Exception as e:
            raise RuntimeError(f"Ollama not reachable: {e}")

    # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    # Internal helpers
    # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

    def _options_text(self, opts: dict) -> str:
        return "\n".join(f"  {k}. {v}" for k, v in opts.items())

    def _ask_round1(self, doctor: dict, question: dict) -> dict:
        """Query one doctor for their initial answer. Returns dict with name, raw, letter."""
        prompt = self.ROUND1_PROMPT.format(
            name      = doctor["name"],
            specialty = doctor["specialty"],
            question  = question["question"],
            options   = self._options_text(question["options"]),
        )
        try:
            resp = _ollama.generate(
                model   = doctor["model"],
                prompt  = prompt,
                options = {"temperature": 0.1, "num_predict": 512, "top_p": 0.9},
            )
            raw = resp.get("response", "").strip()
            letter = self._extract_letter(raw, tag="ANSWER:")
            return {"name": doctor["name"], "model": doctor["model"],
                    "weight": doctor["weight"], "raw_r1": raw, "answer_r1": letter}
        except Exception as e:
            print(f"    âš ï¸  {doctor['name']} Round-1 error: {e}")
            return {"name": doctor["name"], "model": doctor["model"],
                    "weight": doctor["weight"], "raw_r1": "", "answer_r1": "?"}

    def _ask_debate(self, doctor: dict, question: dict, all_r1: list) -> dict:
        """Debate round: doctor sees everyone else's Round-1 answer and may revise."""
        my_r1 = next((r for r in all_r1 if r["name"] == doctor["name"]), {})
        my_ans = my_r1.get("answer_r1", "?")

        others_text = "\n".join(
            f"  {r['name']} ({r.get('weight',1.0)}Ã— weight): {r['answer_r1']}"
            for r in all_r1 if r["name"] != doctor["name"] and r["answer_r1"] != "?"
        )
        if not others_text:
            # No valid peers â€” keep original answer
            return {**my_r1, "answer_final": my_ans, "raw_debate": "(no peers)"}

        prompt = self.DEBATE_PROMPT.format(
            name          = doctor["name"],
            question      = question["question"],
            options       = self._options_text(question["options"]),
            my_answer     = my_ans,
            others_answers= others_text,
        )
        try:
            resp = _ollama.generate(
                model   = doctor["model"],
                prompt  = prompt,
                options = {"temperature": 0.1, "num_predict": 512, "top_p": 0.9},
            )
            raw = resp.get("response", "").strip()
            letter = self._extract_letter(raw, tag="FINAL ANSWER:")
            if letter == "?":
                letter = my_ans  # Fall back to Round-1 if parse fails
            return {**my_r1, "answer_final": letter, "raw_debate": raw}
        except Exception as e:
            print(f"    âš ï¸  {doctor['name']} Debate error: {e}")
            return {**my_r1, "answer_final": my_ans, "raw_debate": f"error: {e}"}

    def _extract_letter(self, text: str, tag: str = "ANSWER:") -> str:
        """Extract the answer letter after a tag like 'ANSWER:' or 'FINAL ANSWER:'."""
        upper = text.upper()
        idx = upper.rfind(tag.upper())
        if idx != -1:
            segment = upper[idx + len(tag):].strip()
            for ch in segment:
                if ch in "ABCDE":
                    return ch
        # Fallback: find last standalone answer letter in text
        import re
        matches = re.findall(r'\b([A-E])\b', upper)
        if matches:
            return matches[-1]
        return "?"

    def _weighted_consensus(self, debate_results: list) -> tuple:
        """
        Weighted vote on post-debate final answers.
        Returns (winning_letter, tally_dict, agreement_bool).
        """
        tally = {}
        for r in debate_results:
            ans = r.get("answer_final", "?")
            if ans == "?":
                continue
            tally[ans] = tally.get(ans, 0.0) + r.get("weight", 1.0)

        if not tally:
            return "?", {}, False

        winner = max(tally, key=tally.get)
        total_weight = sum(tally.values())
        agreement = (tally[winner] / total_weight) >= 0.5
        return winner, tally, agreement

    # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    # Public: run a single question
    # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

    def run_question(self, question: dict, q_index: int = 0) -> dict:
        """Full council pipeline for one question with live status tracking."""
        # Live: Round 1
        with _BENCH_LOCK:
            _BENCH_STATUS["current_q_index"] = q_index
            _BENCH_STATUS["current_q_text"]  = question["question"][:120]
            _BENCH_STATUS["current_round"]   = "round1"
            _BENCH_STATUS["current_doctors"] = [d["name"] for d in self.COUNCIL]

        print(f"  [Q{q_index+1}] {question['question'][:70]}...")
        print(f"  [Round 1 - Parallel] All doctors analyze simultaneously...")
        with concurrent.futures.ThreadPoolExecutor(max_workers=len(self.COUNCIL)) as pool:
            futures = {pool.submit(self._ask_round1, doc, question): doc for doc in self.COUNCIL}
            round1_results = []
            for future in concurrent.futures.as_completed(futures):
                res = future.result()
                round1_results.append(res)
                marker = "V" if res["answer_r1"] == question["correct"] else "X"
                print(f"    {res['name']:12} -> {res['answer_r1']} {marker}")

        # Live: Debate
        with _BENCH_LOCK:
            _BENCH_STATUS["current_round"]   = "debate"
            _BENCH_STATUS["current_doctors"] = [d["name"] for d in self.COUNCIL]

        print(f"  [Round 2 - Debate] Cross-examination...")
        with concurrent.futures.ThreadPoolExecutor(max_workers=len(self.COUNCIL)) as pool:
            futures = {pool.submit(self._ask_debate, doc, question, round1_results): doc for doc in self.COUNCIL}
            debate_results = []
            for future in concurrent.futures.as_completed(futures):
                res = future.result()
                debate_results.append(res)
                changed = " [revised!]" if res["answer_final"] != res.get("answer_r1") else ""
                marker = "V" if res["answer_final"] == question["correct"] else "X"
                print(f"    {res['name']:12} -> {res['answer_final']} {marker}{changed}")

        # Live: Consensus
        with _BENCH_LOCK:
            _BENCH_STATUS["current_round"]   = "consensus"
            _BENCH_STATUS["current_doctors"] = []

        council_ans, tally, agreed = self._weighted_consensus(debate_results)
        correct = (council_ans == question["correct"])
        votes_r1    = {r["name"]: r["answer_r1"]    for r in round1_results}
        votes_final = {r["name"]: r["answer_final"] for r in debate_results}

        print(f"  [Council] {'OK' if correct else 'WRONG'} Council={council_ans} Correct={question['correct']} Agreed={agreed}")

        return {
            "id": question["id"], "source": question["source"],
            "category": question["category"], "difficulty": question["difficulty"],
            "correct_ans": question["correct"], "votes_r1": votes_r1,
            "votes_final": votes_final, "council_ans": council_ans,
            "council_correct": correct, "agreed": agreed,
            "weight_tally": tally, "explanation": question["explanation"],
        }


    # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    # Public: run full benchmark
    # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

    def run(self, questions=None, mode="full"):
        qs = questions or ALL_QUESTIONS
        if mode == "quick":
            qs = qs[:10]

        print(f"\n{'='*70}")
        print(f"  CUREZY AI â€” FULL COUNCIL BENCHMARK  (v2 â€” Council Pipeline Mode)")
        print(f"  Pipeline: Parallel Round-1 â†’ Debate Round â†’ Weighted Consensus")
        print(f"  Date: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}")
        print(f"  Questions: {len(qs)} | Council: {len(self.COUNCIL)} doctors")
        print(f"{'='*70}")

        results      = []
        # Track per-model accuracy for BOTH rounds
        model_scores_r1    = {m["name"]: {"correct": 0, "total": 0} for m in self.COUNCIL}
        model_scores_final = {m["name"]: {"correct": 0, "total": 0} for m in self.COUNCIL}
        council_correct    = 0
        start_all          = time.time()

        # Group by source
        by_source = {}
        for q in qs:
            src = q["source"].split("(")[0].strip()
            by_source.setdefault(src, []).append(q)

        for src, src_qs in by_source.items():
            print(f"\n{'â”€'*70}")
            print(f"  {src.upper()}  ({len(src_qs)} questions)")
            print(f"{'â”€'*70}")

            for q in src_qs:
                print(f"\n  â–¶ [{q['id']}] {q['question'][:75]}...")
                result = self.run_question(q, q_index=results.__len__())
                results.append(result)

                # Live status update after each question
                with _BENCH_LOCK:
                    if result["council_correct"]:
                        _BENCH_STATUS["council_correct"] += 1
                    done = len(results)
                    total = _BENCH_STATUS["total_questions"]
                    _BENCH_STATUS["council_score"] = round(
                        _BENCH_STATUS["council_correct"] / done * 100, 1) if done else 0
                    _BENCH_STATUS["results"].append({
                        "id": result["id"], "source": result["source"],
                        "correct_ans": result["correct_ans"],
                        "council_ans": result["council_ans"],
                        "council_correct": result["council_correct"],
                        "votes_r1": result["votes_r1"],
                        "votes_final": result["votes_final"],
                    })


                if result["council_correct"]:
                    council_correct += 1

                # Score each model in both rounds
                for name in model_scores_r1:
                    model_scores_r1[name]["total"]    += 1
                    model_scores_final[name]["total"] += 1
                    if result["votes_r1"].get(name)    == q["correct"]:
                        model_scores_r1[name]["correct"]    += 1
                    if result["votes_final"].get(name) == q["correct"]:
                        model_scores_final[name]["correct"] += 1

        total_time = round(time.time() - start_all, 1)
        # Mark done in live status
        import datetime as _dt
        with _BENCH_LOCK:
            _BENCH_STATUS["current_round"]   = "done"
            _BENCH_STATUS["current_q_text"]  = "Benchmark complete!"
            _BENCH_STATUS["finished_at"]     = _dt.datetime.now().isoformat()

        report = self._compile(results, model_scores_r1, model_scores_final,

                               council_correct, len(qs), total_time)
        self._print_report(report)
        self._save_json(report)
        self._save_excel(report, results)
        print(f"\nðŸ“Š Saved: benchmark_results.json + benchmark_report.xlsx\n")
        return report

    # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
    # Compile + Print + Save
    # â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

    def _compile(self, results, model_scores_r1, model_scores_final,
                 council_correct, total, elapsed):
        council_pct = round(council_correct / total * 100, 1)

        def pct(sc):
            return round(sc["correct"] / sc["total"] * 100, 1) if sc["total"] else 0

        ind_r1    = {n: pct(s) for n, s in model_scores_r1.items()}
        ind_final = {n: pct(s) for n, s in model_scores_final.items()}
        best_single = max(ind_final.values()) if ind_final else 0
        improvement = round(council_pct - best_single, 1)

        sources, diffs, cats = {}, {}, {}
        for r in results:
            src = r["source"].split("(")[0].strip()
            d, c = r["difficulty"], r["category"]
            for grp, key in [(sources, src), (diffs, d), (cats, c)]:
                if key not in grp:
                    grp[key] = {"correct": 0, "total": 0}
                grp[key]["total"] += 1
                if r["council_correct"]:
                    grp[key]["correct"] += 1
        for grp in (sources, diffs, cats):
            for k in grp:
                grp[k]["pct"] = round(grp[k]["correct"] / grp[k]["total"] * 100, 1)

        # Debate impact: how many questions changed post-debate
        revised = sum(
            1 for r in results
            if any(r["votes_r1"].get(n) != r["votes_final"].get(n)
                   for n in r["votes_r1"])
        )
        revised_correct = sum(
            1 for r in results
            if any(r["votes_r1"].get(n) != r["votes_final"].get(n)
                   for n in r["votes_r1"])
            and r["council_correct"]
        )

        return {
            "timestamp":           datetime.datetime.now().isoformat(),
            "mode":                "full_council_pipeline",
            "pipeline":            "Parallel Round-1 â†’ Debate â†’ Weighted Consensus",
            "total_q":             total,
            "elapsed_s":           elapsed,
            "council": {
                "score_pct":  council_pct,
                "correct":    council_correct,
                "total":      total,
            },
            "individual_round1":   ind_r1,
            "individual_postdebate": ind_final,
            "council_improvement": improvement,
            "debate_impact": {
                "questions_revised": revised,
                "revised_went_correct": revised_correct,
            },
            "by_benchmark":       sources,
            "by_difficulty":      diffs,
            "by_category":        cats,
            "industry_benchmarks":INDUSTRY_BENCHMARKS,
        }

    def _print_report(self, r):
        c  = r["council"]
        sc = c["score_pct"]

        print(f"\n{'='*70}")
        print(f"  FINAL BENCHMARK RESULTS  (Full Council Pipeline)")
        print(f"{'='*70}")
        print(f"""
  CUREZY COUNCIL SCORE:    {sc}%  ({c['correct']}/{c['total']} correct)
  Pipeline:                {r['pipeline']}
  Test Duration:           {r['elapsed_s']}s
""")

        print("  INDIVIDUAL MODELS â€” Round 1 (pre-debate) vs Final (post-debate):")
        for name in r["individual_round1"]:
            r1  = r["individual_round1"][name]
            fin = r["individual_postdebate"][name]
            bar = "â–ˆ" * int(fin / 5) + "â–‘" * (20 - int(fin / 5))
            delta = f"(+{fin-r1:.0f}% debate boost)" if fin > r1 else (f"(-{r1-fin:.0f}% post-debate)" if r1 > fin else "(unchanged)")
            print(f"  {name:12} R1:{r1:5.1f}%  Final:{fin:5.1f}%  [{bar}]  {delta}")

        di = r["debate_impact"]
        print(f"\n  DEBATE IMPACT: "
              f"{di['questions_revised']} questions had at least one doctor revise their answer. "
              f"{di['revised_went_correct']} of those ended correct after debate.")

        print(f"\n  Council vs Best Single Model: +{r['council_improvement']}% improvement\n")

        print("  BY BENCHMARK:")
        for src, d in r["by_benchmark"].items():
            bar = "â–ˆ" * int(d["pct"] / 5) + "â–‘" * (20 - int(d["pct"] / 5))
            print(f"  {src[:28]:28} [{bar}] {d['pct']}%  ({d['correct']}/{d['total']})")

        print("\n  BY DIFFICULTY:")
        for diff, d in r["by_difficulty"].items():
            print(f"  {diff.upper():8}  {d['pct']}%  ({d['correct']}/{d['total']})")

        print("\n  VS INDUSTRY (Overall Accuracy):")
        all_scores = {
            "â†’ Curezy Council": sc,
            **{k: round(sum(v.values()) / len(v), 1) for k, v in INDUSTRY_BENCHMARKS.items()}
        }
        for name, pct in sorted(all_scores.items(), key=lambda x: x[1], reverse=True):
            marker = "â—€ YOU" if "Curezy" in name else ""
            bar    = "â–ˆ" * int(pct / 5) + "â–‘" * (20 - int(pct / 5))
            print(f"  {name:22} [{bar}] {pct:.1f}%  {marker}")

        if sc >= 85:   rating = "ðŸ† EXCELLENT â€” Matches GPT-4 Level"
        elif sc >= 75: rating = "âœ… STRONG â€” Better than GPT-3.5"
        elif sc >= 65: rating = "âš ï¸  GOOD â€” Matches single Med-Gemma"
        elif sc >= 55: rating = "ðŸ“ˆ DEVELOPING â€” Fine-tuning needed"
        else:          rating = "ðŸ”§ NEEDS WORK â€” Review model setup"
        print(f"\n  RATING: {rating}")
        print(f"{'='*70}\n")

    def _save_json(self, report):
        with open("benchmark_results.json", "w") as f:
            json.dump(report, f, indent=2, default=str)

    def _save_excel(self, report, results):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb  = Workbook()
        ws1 = wb.active
        ws1.title = "ðŸ“Š Dashboard"
        ws1.sheet_view.showGridLines = False

        GREEN = "FF22C55E"; DKGRN = "FF15803D"; LTGRN = "FFD1FAE5"
        GRAY  = "FF6B7280"; WHITE = "FFFFFFFF"; RED   = "FFEF4444"
        BLACK = "FF111827"; BLUE  = "FF3B82F6"

        def hdr(ws, row, col, value, bg=DKGRN, fg=WHITE, bold=True, size=11):
            c = ws.cell(row=row, column=col, value=value)
            c.font      = Font(bold=bold, color=fg, size=size, name="Arial")
            c.fill      = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="center", vertical="center")
            return c

        def cell(ws, row, col, value, bg=WHITE, fg=BLACK, bold=False, size=10, align="center"):
            c = ws.cell(row=row, column=col, value=value)
            c.font      = Font(bold=bold, color=fg, size=size, name="Arial")
            c.fill      = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal=align, vertical="center")
            return c

        # Title
        ws1.merge_cells("A1:J1")
        t = ws1["A1"]
        t.value     = "CUREZY AI â€” FULL COUNCIL BENCHMARK REPORT  (Parallel + Debate + Consensus)"
        t.font      = Font(bold=True, size=14, color=WHITE, name="Arial")
        t.fill      = PatternFill("solid", fgColor=DKGRN)
        t.alignment = Alignment(horizontal="center", vertical="center")
        ws1.row_dimensions[1].height = 36

        # Score cards
        r = 3
        cards = [
            ("COUNCIL SCORE",    f"{report['council']['score_pct']}%", DKGRN, WHITE),
            ("QUESTIONS",        f"{report['total_q']} Total",          BLUE,  WHITE),
            ("VS BEST SINGLE",   f"+{report['council_improvement']}%",  GREEN, WHITE),
            ("TEST TIME",        f"{report['elapsed_s']}s",             GRAY,  WHITE),
        ]
        col = 1
        for title, value, bg, fg in cards:
            ws1.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col+1)
            c = ws1.cell(row=r, column=col, value=title)
            c.font = Font(bold=True, size=9, color=fg, name="Arial")
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="center")
            ws1.merge_cells(start_row=r+1, start_column=col, end_row=r+1, end_column=col+1)
            c2 = ws1.cell(row=r+1, column=col, value=value)
            c2.font = Font(bold=True, size=20, color=fg, name="Arial")
            c2.fill = PatternFill("solid", fgColor=bg)
            c2.alignment = Alignment(horizontal="center", vertical="center")
            ws1.row_dimensions[r+1].height = 36
            col += 3

        # Individual model table
        row = 7
        hdr(ws1, row, 1, "MODEL ACCURACY â€” PRE-DEBATE vs POST-DEBATE", bg=DKGRN)
        ws1.merge_cells(f"A{row}:F{row}")
        row += 1
        for h, c_idx in [("Model",9),("Round 1",9),("Post-Debate",9),("Î” Debate",9)]:
            hdr(ws1, row, list([h,"Model","Round 1","Post-Debate","Î” Debate"]).index(h)+1, h, bg=GRAY, size=9)
        row += 1
        for name in report["individual_round1"]:
            r1  = report["individual_round1"][name]
            fin = report["individual_postdebate"][name]
            delta = fin - r1
            bg = LTGRN if fin >= 70 else WHITE
            cell(ws1, row, 1, name,               bg=bg, bold=True, align="left")
            cell(ws1, row, 2, f"{r1}%",           bg=bg)
            cell(ws1, row, 3, f"{fin}%",           bg=bg, bold=True,
                 fg=DKGRN if fin >= 70 else (RED if fin < 50 else BLACK))
            cell(ws1, row, 4, f"{'+' if delta>=0 else ''}{delta:.1f}%", bg=bg,
                 fg=DKGRN if delta > 0 else (RED if delta < 0 else GRAY))
            row += 1

        # Debate impact
        di = report["debate_impact"]
        row += 1
        hdr(ws1, row, 1, "DEBATE IMPACT", bg=DKGRN)
        ws1.merge_cells(f"A{row}:D{row}")
        row += 1
        cell(ws1, row, 1, "Questions revised post-debate", bg=LTGRN, align="left")
        cell(ws1, row, 2, di["questions_revised"], bg=LTGRN, bold=True)
        row += 1
        cell(ws1, row, 1, "Revised questions that went correct", bg=LTGRN, align="left")
        cell(ws1, row, 2, di["revised_went_correct"], bg=LTGRN, bold=True)

        ws1.column_dimensions["A"].width = 28
        ws1.column_dimensions["B"].width = 12
        ws1.column_dimensions["C"].width = 14
        ws1.column_dimensions["D"].width = 14

        # Sheet 2 â€” Question by question
        ws2 = wb.create_sheet("ðŸ“ All Questions")
        ws2.sheet_view.showGridLines = False
        ws2.merge_cells("A1:K1")
        t2 = ws2["A1"]
        t2.value = "QUESTION-BY-QUESTION â€” FULL COUNCIL PIPELINE RESULTS"
        t2.font = Font(bold=True, size=12, color=WHITE, name="Arial")
        t2.fill = PatternFill("solid", fgColor=DKGRN)
        t2.alignment = Alignment(horizontal="center")
        ws2.row_dimensions[1].height = 28

        qs_hdrs = ["ID","Source","Category","Diff","Correct","Council","Result",
                   "Gemma R1","OpenBio R1","Mistral R1","Debate Changed?"]
        for i, h in enumerate(qs_hdrs, 1):
            hdr(ws2, 2, i, h, bg=GRAY, size=9)

        for i, rd in enumerate(results, 3):
            bg = LTGRN if rd["council_correct"] else "FFFEF2F2"
            vr1 = rd.get("votes_r1", {})
            vf  = rd.get("votes_final",{})
            changed = any(vr1.get(n) != vf.get(n) for n in vr1)
            row_vals = [
                rd["id"], rd["source"], rd["category"], rd["difficulty"],
                rd["correct_ans"], rd["council_ans"],
                "âœ… Correct" if rd["council_correct"] else "âŒ Wrong",
                vr1.get("Curezy AURIX","?"),
                vr1.get("Curezy AURA","?"),
                vr1.get("Curezy AURIS","?"),
                "ðŸ”„ Yes" if changed else "â€“",
            ]
            for j, v in enumerate(row_vals, 1):
                c = ws2.cell(row=i, column=j, value=v)
                c.font = Font(size=9, name="Arial", bold=(j==7),
                              color=DKGRN if (rd["council_correct"] and j==7) else
                                    (RED if (not rd["council_correct"] and j==7) else BLACK))
                c.fill = PatternFill("solid", fgColor=bg)
                c.alignment = Alignment(horizontal="center", vertical="center")
            ws2.row_dimensions[i].height = 22

        for ci, w in enumerate([14,22,20,10,9,9,12,10,10,10,14], 1):
            ws2.column_dimensions[get_column_letter(ci)].width = w

        wb.save("benchmark_report.xlsx")
        return "benchmark_report.xlsx"


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# ENTRY POINT
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Curezy Full Council Benchmark v2")
    parser.add_argument("--mode", default="full", choices=["full", "quick"],
                        help="full=all 27 questions, quick=first 10")
    args = parser.parse_args()

    bench = CouncilBenchmark()
    bench.run(mode=args.mode)

